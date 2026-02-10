import os
import json
import hmac
import hashlib
import base64
import uuid
import random
import time
import io
import csv

from flask import Flask, request, Response, redirect, url_for, session, render_template_string
from redis import Redis

from logger import log
from tasks import process_message

from openpyxl import load_workbook


# -----------------------
# ENV / REDIS
# -----------------------
API_KEY = os.getenv("API_KEY")
DEBUG_MODE = os.getenv("DEBUG_MODE", "false").lower() == "true"
LOG_FILE = "/tmp/log.txt"

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
APP_SECRET_KEY = os.getenv("APP_SECRET_KEY", "")

SERVER = os.getenv("SERVER")
REDIS_URL = os.getenv("REDIS_URL")
redis_conn = Redis.from_url(REDIS_URL)

CONFIG_KEY = "config:autoreply"

# Numlist keys
NL_META_KEY = "nl:meta"              # json meta
NL_POOL_LIST = "nl:pool"             # Redis LIST of JSON records (remaining)
NL_ARCHIVE_LIST = "nl:archive"       # optional: consumed history
NL_MESSAGE_KEY = "nl:message"        # message template (UI)
NL_TYPE_KEY = "nl:type"              # sms|mms (UI)

BATCH_INDEX = "nl:batch:index"       # incr counter
BATCH_META_PREFIX = "nl:batch:meta:" # +id -> json
BATCH_ITEMS_PREFIX = "nl:batch:items:"  # +id -> LIST of JSON records


app = Flask(__name__)
app.secret_key = APP_SECRET_KEY or os.urandom(32)


# -----------------------
# AUTH
# -----------------------
def _is_logged_in():
    return session.get("admin_logged_in") is True


def _require_login():
    if not _is_logged_in():
        return redirect(url_for("admin_login"))
    return None


# -----------------------
# CONFIG (autoreply)
# -----------------------
def _get_config_defaults():
    return {
        "reply_mode": 2,
        "step0_type": "sms",
        "step1_type": "sms",
        "step0_text": "",
        "step1_text": "",
    }


def load_config():
    raw = redis_conn.get(CONFIG_KEY)
    defaults = _get_config_defaults()
    if not raw:
        return defaults
    try:
        cfg = json.loads(raw.decode("utf-8"))
        if not isinstance(cfg, dict):
            return defaults
        defaults.update(cfg)
        defaults["reply_mode"] = 1 if int(defaults.get("reply_mode", 2)) == 1 else 2
        if defaults.get("step0_type") not in ("sms", "mms"):
            defaults["step0_type"] = "sms"
        if defaults.get("step1_type") not in ("sms", "mms"):
            defaults["step1_type"] = "sms"
        defaults["step0_text"] = str(defaults.get("step0_text") or "")
        defaults["step1_text"] = str(defaults.get("step1_text") or "")
        return defaults
    except Exception:
        return defaults


def save_config(cfg: dict):
    redis_conn.set(CONFIG_KEY, json.dumps(cfg, ensure_ascii=False))


# -----------------------
# GATEWAY DEVICES
# -----------------------
def _redis_int(key: str) -> int:
    try:
        return int(redis_conn.get(key) or 0)
    except Exception:
        return 0


def _device_stats(device_id: str):
    base = f"stats:device:{device_id}:"
    return {
        "device_id": device_id,
        "received": _redis_int(base + "received"),
        "sent": _redis_int(base + "sent"),
        "errors": _redis_int(base + "errors"),
        "cycle": _redis_int(f"cycle:device:{device_id}:index"),
        "cycle_sent": _redis_int(f"cycle:device:{device_id}:sent"),
        "cycle_received": _redis_int(f"cycle:device:{device_id}:received"),
    }


def fetch_gateway_devices():
    import requests
    if not SERVER or not API_KEY:
        return []
    url = f"{SERVER}/services/get-devices.php"
    try:
        r = requests.get(url, params={"key": API_KEY}, timeout=12)
        data = r.json()
        if not data.get("success"):
            return []
        devices = (data.get("data") or {}).get("devices") or []
        return devices
    except Exception as e:
        log(f"❌ fetch_gateway_devices error: {e}")
        return []


# -----------------------
# NUM LIST PARSING
# -----------------------
def _norm_col(name: str) -> str:
    return (name or "").strip().lower()


def _pick_number_column(columns):
    if not columns:
        return None
    candidates = {"number", "num", "phone", "telephone", "tel", "mobile", "msisdn", "numero", "numéro"}
    for c in columns:
        if _norm_col(c) in candidates:
            return c
    return columns[0]


def _read_csv(file_bytes: bytes):
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        raise Exception("Encodage CSV non supporté")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return [], []

    header = rows[0]
    data_rows = rows[1:] if any(h.strip() for h in header) else rows

    if not any(h.strip() for h in header):
        max_len = max(len(r) for r in rows)
        header = [f"col{i+1}" for i in range(max_len)]
        data_rows = rows

    cleaned = []
    for r in data_rows:
        rr = list(r) + [""] * (len(header) - len(r))
        cleaned.append(rr[:len(header)])

    return header, cleaned


def _read_xlsx(file_bytes: bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else str(v)) for v in row])

    if not rows:
        return [], []

    header = rows[0]
    data_rows = rows[1:]

    if not any(str(h).strip() for h in header):
        max_len = max(len(r) for r in rows)
        header = [f"col{i+1}" for i in range(max_len)]
        data_rows = rows

    header = [str(h).strip() if str(h).strip() else f"col{i+1}" for i, h in enumerate(header)]

    cleaned = []
    for r in data_rows:
        rr = list(r) + [""] * (len(header) - len(r))
        cleaned.append(rr[:len(header)])

    return header, cleaned


def _dedupe_header(header):
    seen = {}
    final_header = []
    for h in header:
        base = (h or "").strip() or "col"
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
        final_header.append(base)
    return final_header


def _build_records(header, rows, number_col):
    idx = header.index(number_col)
    records = []
    for r in rows:
        if idx >= len(r):
            continue
        number = str(r[idx]).strip()
        if not number:
            continue
        rec = {}
        for i, col in enumerate(header):
            rec[col] = str(r[i]).strip() if i < len(r) else ""
        records.append(rec)
    return records


def _nl_remaining_count():
    try:
        return int(redis_conn.llen(NL_POOL_LIST) or 0)
    except Exception:
        return 0


def _load_nl_meta():
    raw = redis_conn.get(NL_META_KEY)
    if not raw:
        return None
    try:
        return json.loads(raw.decode("utf-8"))
    except Exception:
        return None


def _save_nl_meta(meta: dict):
    redis_conn.set(NL_META_KEY, json.dumps(meta, ensure_ascii=False))


def _load_message_draft():
    msg = (redis_conn.get(NL_MESSAGE_KEY) or b"").decode("utf-8", errors="ignore")
    msg_type = (redis_conn.get(NL_TYPE_KEY) or b"sms").decode("utf-8", errors="ignore")
    if msg_type not in ("sms", "mms"):
        msg_type = "sms"
    return msg, msg_type


def _save_message_draft(message: str, msg_type: str):
    redis_conn.set(NL_MESSAGE_KEY, message or "")
    redis_conn.set(NL_TYPE_KEY, msg_type if msg_type in ("sms", "mms") else "sms")


def _template_vars_from_meta(nl_meta):
    if not nl_meta:
        return []
    return list(nl_meta.get("variables") or [])


# -----------------------
# BATCH RESERVATION (consume from pool)
# -----------------------
def _reserve_from_pool(count: int):
    """
    Prend 'count' éléments du pool (remaining) et les retourne (liste de dict).
    ✅ Consommation réelle : on retire du pool.
    """
    if count <= 0:
        return []

    items = []
    pipe = redis_conn.pipeline()
    for _ in range(count):
        pipe.lpop(NL_POOL_LIST)
    raw_items = pipe.execute()

    for raw in raw_items:
        if not raw:
            continue
        try:
            items.append(json.loads(raw.decode("utf-8")))
        except Exception:
            continue
    return items


def _create_batch(selected_device_ids, per_device: int):
    selected_device_ids = [str(x) for x in (selected_device_ids or []) if str(x).strip()]
    per_device = int(per_device or 0)
    if per_device < 0:
        per_device = 0

    nb_devices = len(selected_device_ids)
    total = per_device * nb_devices
    if total <= 0:
        return None, "Total à 0"

    remaining = _nl_remaining_count()
    if remaining <= 0:
        return None, "Numlist vide"

    # si pas assez, on prend ce qu’on peut
    to_take = min(total, remaining)

    batch_id = str(redis_conn.incr(BATCH_INDEX))
    reserved = _reserve_from_pool(to_take)

    # archive (optionnel) -> garder trace de ce qui a été consommé
    if reserved:
        pipe = redis_conn.pipeline()
        for rec in reserved:
            pipe.rpush(NL_ARCHIVE_LIST, json.dumps(rec, ensure_ascii=False))
        pipe.execute()

    # stock batch items
    pipe = redis_conn.pipeline()
    for rec in reserved:
        pipe.rpush(BATCH_ITEMS_PREFIX + batch_id, json.dumps(rec, ensure_ascii=False))
    pipe.execute()

    meta = {
        "batch_id": batch_id,
        "created_at": int(time.time()),
        "devices": selected_device_ids,
        "per_device": per_device,
        "requested_total": total,
        "taken_total": len(reserved),
        "remaining_after": _nl_remaining_count(),
    }
    redis_conn.set(BATCH_META_PREFIX + batch_id, json.dumps(meta, ensure_ascii=False))

    return meta, None


def _list_last_batches(limit=10):
    # on prend les derniers batch_ids via incr index
    idx = _redis_int(BATCH_INDEX)
    out = []
    for i in range(idx, max(0, idx - limit), -1):
        raw = redis_conn.get(BATCH_META_PREFIX + str(i))
        if not raw:
            continue
        try:
            out.append(json.loads(raw.decode("utf-8")))
        except Exception:
            continue
    return out


def _load_batch_items(batch_id: str, limit=50):
    raw_items = redis_conn.lrange(BATCH_ITEMS_PREFIX + str(batch_id), 0, max(0, limit - 1))
    items = []
    for raw in raw_items:
        try:
            items.append(json.loads(raw.decode("utf-8")))
        except Exception:
            continue
    return items


# -----------------------
# ROUTES: LOGIN
# -----------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        pwd = (request.form.get("password") or "").strip()
        if ADMIN_PASSWORD and pwd == ADMIN_PASSWORD:
            session["admin_logged_in"] = True
            return redirect(url_for("admin_settings"))
        return Response("Mot de passe incorrect", status=401, mimetype="text/plain")

    return render_template_string("""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Login</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    :root{--bg:#070a12;--card:#121a2a;--line:#22304a;--txt:#e8eefc;--muted:#8aa0c7;--btn:#2d6cdf;}
    body{margin:0;background:linear-gradient(180deg,#070a12 0%, #0b0f1a 100%);color:var(--txt);font-family:Arial;}
    .wrap{max-width:520px;margin:0 auto;padding:22px;}
    .card{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:18px;margin-top:34px;}
    label{display:block;font-size:12px;color:var(--muted);margin-bottom:6px}
    input{
      width:100%;box-sizing:border-box;background:#0e1626;border:1px solid var(--line);
      color:var(--txt);padding:12px;border-radius:12px;outline:none;
    }
    .btn{background:var(--btn);border:0;color:white;padding:12px 14px;border-radius:12px;cursor:pointer;font-weight:800;margin-top:12px;width:100%}
    h2{margin:0 0 8px 0}
    .muted{color:var(--muted);font-size:12px}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h2>Connexion</h2>
      <div class="muted">Accès au panneau</div>
      <form method="post" style="margin-top:14px">
        <label>Mot de passe</label>
        <input type="password" name="password" autocomplete="current-password">
        <button class="btn" type="submit">Se connecter</button>
      </form>
    </div>
  </div>
</body>
</html>
""")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/admin", methods=["GET"])
def admin_home():
    guard = _require_login()
    if guard:
        return guard
    return redirect(url_for("admin_settings"))


# -----------------------
# ROUTES: NUM LIST
# -----------------------
@app.route("/admin/nl/clear", methods=["GET"])
def admin_nl_clear():
    guard = _require_login()
    if guard:
        return guard

    # clear pool + meta + message draft
    redis_conn.delete(NL_META_KEY)
    redis_conn.delete(NL_POOL_LIST)
    # on ne touche pas l'archive ni les batchs
    return redirect(url_for("admin_settings"))


@app.route("/admin/nl/upload", methods=["POST"])
def admin_nl_upload():
    guard = _require_login()
    if guard:
        return guard

    files = request.files.getlist("files")
    if not files:
        return Response("Fichier manquant", status=400)

    all_records = []
    all_columns = None
    number_col_global = None

    try:
        for f in files:
            filename = (f.filename or "").lower().strip()
            file_bytes = f.read()
            if not file_bytes:
                continue

            if filename.endswith(".csv"):
                header, rows = _read_csv(file_bytes)
            elif filename.endswith(".xlsx"):
                header, rows = _read_xlsx(file_bytes)
            else:
                continue

            header = _dedupe_header(header)
            number_col = _pick_number_column(header)
            if not number_col:
                continue

            records = _build_records(header, rows, number_col)

            if all_columns is None:
                all_columns = header
                number_col_global = number_col
            else:
                for c in header:
                    if c not in all_columns:
                        all_columns.append(c)

            # normalise records sur all_columns
            for rec in records:
                for c in all_columns:
                    rec.setdefault(c, "")
                all_records.append(rec)

        if not all_records or not all_columns:
            return Response("Aucun numéro importé", status=400)

        if number_col_global not in all_columns:
            number_col_global = _pick_number_column(all_columns)

        # push into pool (remaining) WITHOUT clearing existing (tu peux importer plusieurs fois)
        pipe = redis_conn.pipeline()
        for rec in all_records:
            pipe.rpush(NL_POOL_LIST, json.dumps(rec, ensure_ascii=False))
        pipe.execute()

        variables = [c for c in all_columns if c != number_col_global]
        meta = {
            "columns": all_columns,
            "number_col": number_col_global,
            "variables": variables,
            "updated_at": int(time.time()),
        }
        _save_nl_meta(meta)

        return redirect(url_for("admin_settings"))

    except Exception as e:
        log(f"❌ NL upload error: {e}")
        return Response(f"Erreur import: {e}", status=400)


@app.route("/admin/nl/message", methods=["POST"])
def admin_nl_message():
    guard = _require_login()
    if guard:
        return guard

    message = (request.form.get("nl_message") or "").strip()
    msg_type = (request.form.get("nl_type") or "sms").strip().lower()
    _save_message_draft(message, msg_type)
    return redirect(url_for("admin_settings"))


@app.route("/admin/nl/send", methods=["POST"])
def admin_nl_send():
    """
    'Envoyer' dans l'UI = on prépare un lot consommé du pool et on affiche le payload prêt.
    (Le site NE déclenche pas l'envoi outbound automatiquement.)
    """
    guard = _require_login()
    if guard:
        return guard

    per_device = int(request.form.get("per_device") or 0)
    device_ids = request.form.getlist("device_ids")

    meta, err = _create_batch(device_ids, per_device)
    if err:
        return Response(err, status=400, mimetype="text/plain")

    return redirect(url_for("admin_settings", batch=meta["batch_id"]))


# -----------------------
# ROUTES: SETTINGS / UI
# -----------------------
@app.route("/admin/settings", methods=["GET", "POST"])
def admin_settings():
    guard = _require_login()
    if guard:
        return guard

    cfg = load_config()

    if request.method == "POST" and request.form.get("form_name") == "autoreply":
        reply_mode = int(request.form.get("reply_mode") or 2)
        step0_type = (request.form.get("step0_type") or "sms").strip().lower()
        step1_type = (request.form.get("step1_type") or "sms").strip().lower()
        step0_text = (request.form.get("step0_text") or "").strip()
        step1_text = (request.form.get("step1_text") or "").strip()

        if reply_mode not in (1, 2):
            reply_mode = 2
        if step0_type not in ("sms", "mms"):
            step0_type = "sms"
        if step1_type not in ("sms", "mms"):
            step1_type = "sms"
        if reply_mode == 1:
            step1_text = ""

        cfg.update({
            "reply_mode": reply_mode,
            "step0_type": step0_type,
            "step1_type": step1_type,
            "step0_text": step0_text,
            "step1_text": step1_text,
        })
        save_config(cfg)
        return redirect(url_for("admin_settings"))

    nl_meta = _load_nl_meta()
    remaining = _nl_remaining_count()
    nl_message, nl_type = _load_message_draft()

    # devices from gateway
    gw_devices = fetch_gateway_devices()
    rows = []
    for d in gw_devices:
        did = str(d.get("id"))
        s = _device_stats(did)
        s.update({"name": d.get("name") or "", "model": d.get("model") or ""})
        rows.append(s)

    vars_list = _template_vars_from_meta(nl_meta)

    # last batches + optional selected batch
    batches = _list_last_batches(limit=8)
    selected_batch = request.args.get("batch")
    selected_meta = None
    selected_items = []
    if selected_batch:
        raw = redis_conn.get(BATCH_META_PREFIX + str(selected_batch))
        if raw:
            try:
                selected_meta = json.loads(raw.decode("utf-8"))
            except Exception:
                selected_meta = None
        if selected_meta:
            selected_items = _load_batch_items(selected_batch, limit=25)

    return render_template_string("""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Control</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    :root{
      --bg:#070a12; --card:#121a2a; --line:#22304a; --txt:#e8eefc; --muted:#8aa0c7;
      --btn:#2d6cdf; --good:#24d18f;
    }
    body{margin:0;background:linear-gradient(180deg,#070a12 0%, #0b0f1a 100%);color:var(--txt);font-family:Arial;}
    .wrap{max-width:1200px;margin:0 auto;padding:18px;}
    .top{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;}
    .top h2{margin:0;font-size:18px;}
    a{color:#9bc1ff;text-decoration:none;font-weight:700}
    .card{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:16px;margin-top:12px;}
    .muted{color:var(--muted);font-size:12px}
    .title{font-weight:900;margin-bottom:10px}
    .row{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end}
    label{display:block;font-size:12px;color:var(--muted);margin-bottom:6px}
    textarea, select, input[type="file"], input[type="number"]{
      width:100%;box-sizing:border-box;background:#0e1626;border:1px solid var(--line);
      color:var(--txt);padding:12px;border-radius:12px;outline:none;
    }
    textarea{min-height:90px;resize:vertical}
    select{appearance:none;background-image:
      linear-gradient(45deg,transparent 50%,#9bc1ff 50%),
      linear-gradient(135deg,#9bc1ff 50%,transparent 50%);
      background-position: calc(100% - 18px) calc(50% - 3px), calc(100% - 12px) calc(50% - 3px);
      background-size: 6px 6px, 6px 6px;
      background-repeat:no-repeat;
    }
    .actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:10px}
    .btn{
      height:42px; display:inline-flex; align-items:center; justify-content:center;
      padding:0 16px; border-radius:12px; cursor:pointer; font-weight:900;
      border:1px solid var(--line); text-decoration:none;
    }
    .btn-primary{background:var(--btn); color:white; border:0}
    .btn-secondary{background:#0e1626; color:#cfe0ff}
    .btn-danger{background:#2a1220; color:#ffb6c6; border:1px solid #4a22304a}
    .pill{display:inline-flex;gap:8px;align-items:center;background:#0e1626;border:1px solid var(--line);padding:10px 12px;border-radius:14px}
    .dot{width:9px;height:9px;border-radius:99px;background:var(--good)}
    table{width:100%;border-collapse:collapse}
    th,td{padding:10px;border-bottom:1px solid var(--line);text-align:left;font-size:13px}
    th{color:var(--muted);font-weight:900}
    code{background:#0e1626;padding:2px 6px;border-radius:10px;border:1px solid var(--line)}
    .chips{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}
    .chip{
      border:1px solid var(--line); background:#0e1626; color:#cfe0ff;
      padding:8px 10px; border-radius:999px; cursor:pointer; font-weight:900; font-size:12px;
    }
    .chip:hover{filter:brightness(1.15)}
    .grid2{display:grid;grid-template-columns:1fr;gap:12px}
    @media(min-width:900px){ .grid2{grid-template-columns:1fr 1fr} }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h2>Panneau de contrôle</h2>
      <div style="display:flex;gap:12px;align-items:center">
        <a href="/logs" target="_blank">Logs</a>
        <a href="/admin/logout">Logout</a>
      </div>
    </div>

    <!-- DEVICES -->
    <div class="card">
      <div class="title">Appareils</div>
      <table>
        <thead>
          <tr>
            <th>Device</th>
            <th>Nom / Modèle</th>
            <th>Reçus</th>
            <th>Envoyés</th>
            <th>Erreurs</th>
            <th>Cycle</th>
            <th>Reçus cycle</th>
            <th>Envoyés cycle</th>
          </tr>
        </thead>
        <tbody>
          {% if rows|length == 0 %}
            <tr><td colspan="8" class="muted">Aucun device (vérifie SERVER/API_KEY).</td></tr>
          {% endif %}
          {% for r in rows %}
            <tr>
              <td><div class="pill"><span class="dot"></span><span style="font-weight:900">#{{ r.device_id }}</span></div></td>
              <td><div style="font-weight:900">{{ r.name }}</div><div class="muted">{{ r.model }}</div></td>
              <td>{{ r.received }}</td>
              <td>{{ r.sent }}</td>
              <td>{{ r.errors }}</td>
              <td>{{ r.cycle }}</td>
              <td>{{ r.cycle_received }}</td>
              <td>{{ r.cycle_sent }}</td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <!-- NUMLIST + MESSAGE -->
    <div class="card">
      <div class="title">Numlist + Message</div>

      <div class="row">
        <div class="pill" style="min-width:260px">
          <div>
            <div class="muted">Numéros restants</div>
            <div style="font-size:26px;font-weight:900">{{ remaining }}</div>
          </div>
        </div>

        {% if nl_meta %}
          <div class="pill" style="min-width:260px">
            <div>
              <div class="muted">Colonne numéro</div>
              <div style="font-weight:900"><code>{{ nl_meta.number_col }}</code></div>
            </div>
          </div>
        {% endif %}
      </div>

      <div class="grid2" style="margin-top:12px">
        <div>
          <form method="post" action="/admin/nl/upload" enctype="multipart/form-data">
            <label>Importer des fichiers (.xlsx ou .csv)</label>
            <input type="file" name="files" accept=".xlsx,.csv" multiple required>
            <div class="actions">
              <button class="btn btn-secondary" type="submit">Importer</button>
              <a class="btn btn-danger" href="/admin/nl/clear">Vider</a>
            </div>
          </form>

          {% if vars_list|length > 0 %}
            <div class="muted" style="margin-top:12px">Variables détectées :</div>
            <div class="chips">
              {% for v in vars_list %}
                <div class="chip" onclick="insertVar('{{ v }}')">{{'{{'}}{{ v }}{{'}}'}}</div>
              {% endfor %}
            </div>
          {% else %}
            <div class="muted" style="margin-top:12px">Aucune variable (fichier 1 colonne ou seulement numéro)</div>
          {% endif %}
        </div>

        <div>
          <form method="post" action="/admin/nl/message">
            <div class="row">
              <div style="min-width:220px;flex:1;max-width:320px">
                <label>Type</label>
                <select name="nl_type">
                  <option value="sms" {% if nl_type == 'sms' %}selected{% endif %}>sms</option>
                  <option value="mms" {% if nl_type == 'mms' %}selected{% endif %}>mms</option>
                </select>
              </div>
            </div>

            <div style="margin-top:10px">
              <label>Message</label>
              <textarea id="nl_message" name="nl_message">{{ nl_message }}</textarea>
            </div>

            <div class="actions">
              <button class="btn btn-primary" type="submit">Enregistrer message</button>
            </div>
          </form>
        </div>
      </div>

      <form method="post" action="/admin/nl/send" style="margin-top:14px">
        <div class="row">
          <div style="min-width:260px;max-width:320px;flex:1">
            <label>Nombre par appareil</label>
            <input type="number" min="0" name="per_device" value="0">
          </div>
        </div>

        <div class="muted" style="margin-top:10px">Appareils sélectionnés :</div>
        <div class="chips" style="margin-top:8px">
          {% for r in rows %}
            <label class="chip" style="display:inline-flex;align-items:center;gap:8px">
              <input type="checkbox" name="device_ids" value="{{ r.device_id }}" style="accent-color:#2d6cdf">
              #{{ r.device_id }}
            </label>
          {% endfor %}
        </div>

        <div class="actions" style="margin-top:12px">
          <button class="btn btn-primary" type="submit">Envoyer (prépare le lot)</button>
        </div>

        <div class="muted" style="margin-top:8px">
          Ce bouton prépare un lot (consomme des numéros) et affiche un payload prêt à envoyer via ton gateway.
        </div>
      </form>
    </div>

    <!-- BATCH RESULT -->
    {% if selected_meta %}
      <div class="card">
        <div class="title">Lot #{{ selected_meta.batch_id }}</div>
        <div class="muted">
          Pris: <b>{{ selected_meta.taken_total }}</b> / demandé: {{ selected_meta.requested_total }} •
          Restants: <b>{{ selected_meta.remaining_after }}</b>
        </div>

        <div style="margin-top:10px" class="muted">Payload (extrait 25 lignes max) :</div>
        <div style="margin-top:6px;max-height:260px;overflow:auto;border:1px solid var(--line);border-radius:12px">
          <table>
            <thead>
              <tr>
                <th>number</th>
                <th>data</th>
              </tr>
            </thead>
            <tbody>
              {% for rec in selected_items %}
                <tr>
                  <td>{{ rec.get(nl_meta.number_col, '') if nl_meta else '' }}</td>
                  <td class="muted">{{ rec }}</td>
                </tr>
              {% endfor %}
            </tbody>
          </table>
        </div>

        <div class="muted" style="margin-top:10px">
          Type: <code>{{ nl_type }}</code> • Message: enregistré (variables possible).
        </div>
      </div>
    {% endif %}

    <!-- LAST BATCHES -->
    {% if batches|length > 0 %}
      <div class="card">
        <div class="title">Derniers lots</div>
        <table>
          <thead>
            <tr>
              <th>ID</th>
              <th>Pris</th>
              <th>Demandé</th>
              <th>Appareils</th>
              <th>Restants après</th>
              <th>Voir</th>
            </tr>
          </thead>
          <tbody>
            {% for b in batches %}
              <tr>
                <td>#{{ b.batch_id }}</td>
                <td>{{ b.taken_total }}</td>
                <td>{{ b.requested_total }}</td>
                <td class="muted">{{ b.devices }}</td>
                <td>{{ b.remaining_after }}</td>
                <td><a href="/admin/settings?batch={{ b.batch_id }}">ouvrir</a></td>
              </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    {% endif %}

    <!-- AUTOREPLY (optionnel) -->
    <div class="card">
      <div class="title">Auto-reply</div>
      <form method="post">
        <input type="hidden" name="form_name" value="autoreply">

        <div class="row">
          <div style="min-width:260px;flex:1;max-width:320px">
            <label>Mode</label>
            <select id="reply_mode" name="reply_mode">
              <option value="1" {% if cfg.reply_mode == 1 %}selected{% endif %}>1 réponse (puis stop)</option>
              <option value="2" {% if cfg.reply_mode == 2 %}selected{% endif %}>2 réponses (puis stop)</option>
            </select>
          </div>
        </div>

        <div class="card" style="padding:12px;margin-top:10px">
          <div class="title">Step 1</div>
          <div class="row">
            <div style="min-width:220px;flex:1;max-width:320px">
              <label>Type</label>
              <select name="step0_type">
                <option value="sms" {% if cfg.step0_type == 'sms' %}selected{% endif %}>sms</option>
                <option value="mms" {% if cfg.step0_type == 'mms' %}selected{% endif %}>mms</option>
              </select>
            </div>
          </div>
          <div style="margin-top:10px">
            <label>Message</label>
            <textarea name="step0_text">{{ cfg.step0_text }}</textarea>
          </div>
        </div>

        <div id="step2_block" class="card" style="padding:12px;margin-top:10px">
          <div class="title">Step 2</div>
          <div class="row">
            <div style="min-width:220px;flex:1;max-width:320px">
              <label>Type</label>
              <select name="step1_type">
                <option value="sms" {% if cfg.step1_type == 'sms' %}selected{% endif %}>sms</option>
                <option value="mms" {% if cfg.step1_type == 'mms' %}selected{% endif %}>mms</option>
              </select>
            </div>
          </div>
          <div style="margin-top:10px">
            <label>Message</label>
            <textarea name="step1_text">{{ cfg.step1_text }}</textarea>
          </div>
        </div>

        <div class="actions">
          <button class="btn btn-primary" type="submit">Sauvegarder réponses</button>
        </div>
      </form>
    </div>

  </div>

  <script>
    function insertVar(name){
      const el = document.getElementById("nl_message");
      if(!el) return;
      const token = "{{" + name + "}}";
      const start = el.selectionStart || 0;
      const end = el.selectionEnd || 0;
      const before = el.value.substring(0, start);
      const after = el.value.substring(end);
      el.value = before + token + after;
      el.focus();
      const pos = start + token.length;
      el.setSelectionRange(pos, pos);
    }

    function toggleStep2() {
      const mode = document.getElementById("reply_mode").value;
      const block = document.getElementById("step2_block");
      if (mode === "1") block.style.display = "none";
      else block.style.display = "block";
    }
    document.getElementById("reply_mode").addEventListener("change", toggleStep2);
    toggleStep2();
  </script>
</body>
</html>
""",
        cfg=cfg,
        rows=rows,
        nl_meta=nl_meta,
        remaining=remaining,
        nl_message=nl_message,
        nl_type=nl_type,
        vars_list=vars_list,
        batches=batches,
        selected_meta=selected_meta,
        selected_items=selected_items,
    )


# -----------------------
# WEBHOOK (inchangé)
# -----------------------
@app.route("/sms_auto_reply", methods=["POST"])
def sms_auto_reply():
    request_id = str(uuid.uuid4())[:8]
    log(f"\n📩 [{request_id}] Nouvelle requête POST reçue")

    messages_raw = request.form.get("messages")
    if not messages_raw:
        log(f"[{request_id}] ❌ Champ 'messages' manquant")
        return "messages manquants", 400

    if not DEBUG_MODE:
        signature = request.headers.get("X-SG-SIGNATURE")
        if not signature:
            log(f"[{request_id}] ❌ Signature manquante")
            return "Signature requise", 403

        expected_hash = base64.b64encode(
            hmac.new(API_KEY.encode(), messages_raw.encode(), hashlib.sha256).digest()
        ).decode()

        if signature != expected_hash:
            log(f"[{request_id}] ❌ Signature invalide")
            return "Signature invalide", 403

    try:
        messages = json.loads(messages_raw)
    except json.JSONDecodeError as e:
        log(f"[{request_id}] ❌ JSON invalide : {e}")
        return "Format JSON invalide", 400

    if not isinstance(messages, list):
        return "Liste attendue", 400

    for msg in messages:
        try:
            delay = random.randint(60, 180)
            process_message.apply_async(args=[json.dumps(msg)], countdown=delay)
        except Exception as e:
            log(f"[{request_id}] ❌ Erreur Celery : {e}")

    return "OK", 200


@app.route("/logs")
def logs():
    if not os.path.exists(LOG_FILE):
        return Response("Aucun log", mimetype="text/plain")
    with open(LOG_FILE, "r", encoding="utf-8") as f:
        return Response(f.read(), mimetype="text/plain")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
